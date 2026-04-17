import streamlit as st
import pandas as pd
from sqlalchemy import text
import openpyxl
import re

# ==========================================
# 0. 安全驗證
# ==========================================
def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["passwords"]["admin_password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False
    if "password_correct" not in st.session_state:
        st.title("🔒 營收管理系統 - 登入")
        st.text_input("請輸入存取密碼", type="password", on_change=password_entered, key="password")
        return False
    return st.session_state.get("password_correct", False)

if check_password():
    st.set_page_config(page_title="車聯網營收戰情系統 v20", layout="wide")
    conn = st.connection("postgresql", type="sql")

    # ==========================================
    # 1. 核心讀取與清洗區
    # ==========================================
    def load_data():
        try:
            df = conn.query("SELECT * FROM financials", ttl="0")
            months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
            for m in months:
                if m in df.columns:
                    df[m] = pd.to_numeric(df[m], errors='coerce').fillna(0)
            return df
        except:
            return pd.DataFrame()

    def save_to_supabase(df):
        with conn.session as session:
            session.execute(text("DELETE FROM financials"))
            session.commit()
        df.to_sql('financials', conn.engine, if_exists='append', index=False, chunksize=50, method='multi')

    def clean_currency(val):
        """強制把 $、逗號、括號(負數) 轉成純數字"""
        if pd.isna(val): return 0.0
        val_str = str(val).strip()
        if not val_str or val_str.lower() in ['nan', 'none', 'null']: return 0.0
        val_str = re.sub(r'[^\d\.\-\(\)]', '', val_str)
        if val_str.startswith('(') and val_str.endswith(')'):
            val_str = '-' + val_str[1:-1]
        try:
            return float(val_str)
        except:
            return 0.0

    def process_imported_file(uploaded_file):
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = wb.sheetnames
        target_s = sheet_names[0]
        for s in sheet_names:
            if any(k in s for k in ['營收', '預估', '收支']):
                target_s = s
                break
        sheet = wb[target_s]
        
        color_list = []
        for row in sheet.iter_rows(min_row=4):
            color_id = "無底色"
            for cell in [row[1], row[2], row[19]]: 
                fill = cell.fill
                if fill and hasattr(fill, 'start_color') and fill.start_color:
                    sc = fill.start_color
                    if getattr(sc, 'type', None) == 'rgb' and getattr(sc, 'rgb', None):
                        val = str(sc.rgb)
                        if val != '00000000' and val != '000000': 
                            color_id = f"色碼_{val[-6:].upper()}"
                            break
                    elif getattr(sc, 'type', None) == 'theme':
                        color_id = f"主題色_T{getattr(sc, 'theme', '未知')}_色偏{getattr(sc, 'tint', 0.0)}"
                        break
                    elif getattr(sc, 'type', None) == 'indexed':
                        color_id = f"索引色_{getattr(sc, 'indexed', '未知')}"
                        break
                    else:
                        try:
                            raw = str(getattr(sc, 'rgb', '無RGB資料'))
                            if raw != '00000000' and raw != '無RGB資料':
                                safe_raw = "".join(char for char in raw if char.isalnum())
                                color_id = f"特殊格式_{safe_raw[-6:]}"
                                break
                        except:
                            color_id = "無法解析的特殊顏色"
                            break
            color_list.append(color_id)

        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=target_s, skiprows=2)
        df.columns = [str(c).strip() for c in df.columns]
        
        df['顏色標記'] = color_list[:len(df)]
        df.rename(columns={df.columns[2]: '紀錄類型'}, inplace=True)
        df['專案說明'] = df['專案說明'].replace(r'^\s*$', pd.NA, regex=True).ffill()
        
        cat_col_name = ""
        for c in df.columns:
            if '營收分類' in str(c):
                cat_col_name = c
                break
        
        if cat_col_name:
            df[cat_col_name] = df[cat_col_name].astype(str).str.replace('\n', ' ').str.strip()
            df[cat_col_name] = df[cat_col_name].replace(['nan', 'None', '', '<NA>', 'NaN'], pd.NA)
            df['營收分類'] = df[cat_col_name].ffill().fillna("其他")
        else:
            df['營收分類'] = "其他"
        
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        for m in months:
            if m in df.columns:
                df[m] = df[m].apply(clean_currency)
            else:
                df[m] = 0.0

        fallback_cols = [c for c in df.columns if any(k in str(c) for k in ['小計', '總計', '合計', '實績'])]
        df['temp_sum'] = df[months].sum(axis=1)
        for idx, row in df.iterrows():
            if row['temp_sum'] == 0:
                for f_col in fallback_cols:
                    if f_col in row:
                        try:
                            val_str = str(row[f_col]).replace(',', '').strip()
                            if val_str and val_str.lower() not in ['nan', 'none', '']:
                                val = float(val_str)
                                if val != 0:
                                    df.at[idx, 'Jan'] = val
                                    break
                        except:
                            pass
        df = df.drop(columns=['temp_sum'])
        
        if '說明' not in df.columns: df['說明'] = ""
        target_cols = ['專案說明', '紀錄類型'] + months + ['營收分類', '顏色標記', '說明']
        return df.dropna(subset=['專案說明', '紀錄類型'])[target_cols]

    # ==========================================
    # 2. 側邊欄：匯入機制
    # ==========================================
    with st.sidebar:
        st.header("📂 匯入數據")
        f = st.file_uploader("選擇 Excel", type=["xlsx"])
        if f and st.button("🚀 開始解析並上傳"):
            with st.spinner("強制淨化數字格式與清洗分類..."):
                try:
                    new_df = process_imported_file(f)
                    save_to_supabase(new_df)
                    st.success("匯入成功！所有符號皆已清除。")
                    st.rerun()
                except Exception as e:
                    st.error(f"解析失敗: {e}")
        
        st.divider()
        if st.button("⚠️ 清空資料庫"):
            with conn.session as s:
                s.execute(text("DELETE FROM financials"))
                s.commit()
            st.rerun()

    # ==========================================
    # 3. 介面呈現
    # ==========================================
    st.title("📊 車聯網事業本部 - 專案營收戰情室")
    data = load_data()

    if data.empty:
        st.warning("⚠️ 目前資料庫為空，請從左側邊欄匯入 Excel 檔案。")
    else:
        # --- 全局變數準備 ---
        df_sum = data.copy()
        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        df_sum['年度總計'] = df_sum[months].sum(axis=1)
        
        unique_colors = df_sum['顏色標記'].dropna().unique().tolist()
        if "無底色" not in unique_colors: unique_colors.insert(0, "無底色")
        unique_records = df_sum['紀錄類型'].dropna().unique().tolist()

        # 自動猜測預設選項 (防呆機制)
        def_target_c = [c for c in unique_colors if 'D9D9D9' in c]
        def_est_c = [c for c in unique_colors if 'F2DCDB' in c]
        def_inc = [r for r in unique_records if r in ['收入', '營收', '實績']]
        def_exp = [r for r in unique_records if r in ['支出', '成本']]

        # --- 🚀 全新武器：全局報表綁定面板 ---
        with st.expander("⚙️ 第 1 步：定義報表邏輯 (完全杜絕關鍵字重複問題)", expanded=True):
            st.markdown("##### 🎨 1. 顏色定義 (決定資料落入『目標』還是『預估』)")
            c1, c2 = st.columns(2)
            with c1:
                target_colors = st.multiselect("🎯 勾選代表『原目標』的底色", unique_colors, default=def_target_c)
            with c2:
                est_colors = st.multiselect("🔮 勾選代表『預估』的底色", unique_colors, default=def_est_c)

            st.divider()
            st.markdown("##### 📝 2. 文字定義 (決定資料落入『收入』還是『支出』)")
            st.caption("請直接勾選 Excel 內出現的完整文字。例如：將『收入』勾在左邊，將『收入預估』也勾在左邊。系統會透過上方的『顏色』來區分它是目標還是預估！")
            c3, c4 = st.columns(2)
            with c3:
                income_records = st.multiselect("💰 勾選屬於『收入屬性』的項目", unique_records, default=def_inc)
            with c4:
                expense_records = st.multiselect("💸 勾選屬於『支出屬性』的項目", unique_records, default=def_exp)

        # 產生過濾條件 (精準到小數點等級，絕不重疊)
        is_target_color = df_sum['顏色標記'].isin(target_colors)
        is_est_color = df_sum['顏色標記'].isin(est_colors)
        is_income_text = df_sum['紀錄類型'].isin(income_records)
        is_expense_text = df_sum['紀錄類型'].isin(expense_records)

        tabs = st.tabs(["📈 營收分類總表", "🎴 專案卡片摘要", "📝 原始數據管理"])

        # --- TAB 1: 營收分類總表 ---
        with tabs[0]:
            st.subheader("📋 營收分類戰情總表")
            unique_cats = df_sum['營收分類'].unique().tolist()
            
            summary = pd.DataFrame(index=unique_cats)
            
            # 使用我們剛才在面板設定的精確選項，兩兩交集
            summary['原目標收入'] = df_sum[is_target_color & is_income_text].groupby('營收分類')['年度總計'].sum()
            summary['預估收入'] = df_sum[is_est_color & is_income_text].groupby('營收分類')['年度總計'].sum()
            summary['實際支出'] = df_sum[is_target_color & is_expense_text].groupby('營收分類')['年度總計'].sum()
            summary['預估支出'] = df_sum[is_est_color & is_expense_text].groupby('營收分類')['年度總計'].sum()
            
            summary = summary.fillna(0)

            summary['原毛利'] = summary['原目標收入'] - summary['實際支出']
            summary['預估毛利'] = summary['預估收入'] - summary['預估支出']
            summary['差異'] = summary['預估收入'] - summary['原目標收入']
            summary['毛利率'] = (summary['預估毛利'] / summary['預估收入']).replace([float('inf'), -float('inf')], 0).fillna(0)

            summary = summary.reset_index().rename(columns={'index': '營收分類'})
            display_cols = ['營收分類', '原目標收入', '預估收入', '原毛利', '預估毛利', '差異', '毛利率']
            
            st.dataframe(
                summary[display_cols].style.format({
                    '原目標收入': '{:,.0f}', '預估收入': '{:,.0f}', 
                    '原毛利': '{:,.0f}', '預估毛利': '{:,.0f}', 
                    '差異': '{:,.0f}', '毛利率': '{:.2%}'
                }).map(lambda x: 'color: red' if isinstance(x, (int, float)) and x < 0 else '', subset=['差異', '預估毛利']),
                use_container_width=True,
                hide_index=True
            )

        # --- TAB 2: 專案卡片摘要 ---
        with tabs[1]:
            st.subheader("💡 專案績效一覽表")
            cats = ["全部分類"] + unique_cats
            sel_cat = st.selectbox("篩選營收分類", cats, key="card_filter")
            
            display_data = df_sum if sel_cat == "全部分類" else df_sum[df_sum['營收分類'] == sel_cat]
            projects = display_data['專案說明'].unique()
            
            cols = st.columns(2)
            for idx, proj in enumerate(projects):
                with cols[idx % 2]:
                    p_df = display_data[display_data['專案說明'] == proj]
                    
                    # 卡片同樣受惠於全局面板的精確選擇
                    p_is_target = p_df['顏色標記'].isin(target_colors)
                    p_is_est = p_df['顏色標記'].isin(est_colors)
                    p_is_inc = p_df['紀錄類型'].isin(income_records)
                    p_is_exp = p_df['紀錄類型'].isin(expense_records)
                    
                    target = p_df[p_is_target & p_is_inc]['年度總計'].sum()
                    est_in = p_df[p_is_est & p_is_inc]['年度總計'].sum()
                    est_out = p_df[p_is_est & p_is_exp]['年度總計'].sum()
                    
                    profit = est_in - est_out
                    margin = (profit / est_in) if est_in != 0 else 0
                    
                    with st.container(border=True):
                        st.markdown(f"#### {proj}")
                        cat_name = p_df['營收分類'].iloc[0] if not p_df['營收分類'].empty else '未知'
                        st.caption(f"營收分類：{cat_name}")
                        
                        m1, m2, m3 = st.columns(3)
                        m1.metric("目標營收", f"${target:,.0f}")
                        m2.metric("預估營收", f"${est_in:,.0f}", f"{est_in-target:,.0f}")
                        m3.metric("預估毛利率", f"{margin:.1%}")
                        
                        reach = (est_in / target) if target != 0 else 0
                        st.write(f"**目標達成率: {reach:.1%}**")
                        st.progress(min(max(reach, 0.0), 1.0))
                        
                        with st.expander("查看明細數據"):
                            st.dataframe(p_df[['紀錄類型', '顏色標記', '年度總計']], use_container_width=True)

        # --- TAB 3: 原始數據管理 ---
        with tabs[2]:
            st.info("💡 在此可以直接修改資料庫內的數據，修改後請點擊下方儲存。")
            edited = st.data_editor(data, num_rows="dynamic", use_container_width=True, height=500)
            if st.button("💾 儲存變更", type="primary"):
                save_to_supabase(edited)
                st.success("雲端已更新！")
